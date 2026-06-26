"""SQLite-backed metrics store with CSV (微信视频号) and XLSX (小红书) import."""
from __future__ import annotations

import csv
import io
import re
import sqlite3
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

_SCHEMA = """
CREATE TABLE IF NOT EXISTS video_metrics (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    platform        TEXT NOT NULL,
    title           TEXT NOT NULL,
    platform_id     TEXT,
    publish_date    TEXT NOT NULL,
    content_type    TEXT DEFAULT 'video',
    plays           INTEGER DEFAULT 0,
    likes           INTEGER DEFAULT 0,
    comments        INTEGER DEFAULT 0,
    shares          INTEGER DEFAULT 0,
    followers_gained INTEGER DEFAULT 0,
    completion_rate REAL,
    avg_watch_duration REAL,
    exposure        INTEGER DEFAULT 0,
    cover_click_rate REAL,
    favorites       INTEGER DEFAULT 0,
    danmaku         INTEGER DEFAULT 0,
    forward_count   INTEGER DEFAULT 0,
    extra           TEXT,
    job_id          TEXT,
    used_asset_ids  TEXT,
    imported_at     TEXT NOT NULL,
    UNIQUE(platform, title, publish_date)
);
CREATE INDEX IF NOT EXISTS idx_metrics_platform ON video_metrics(platform);
CREATE INDEX IF NOT EXISTS idx_metrics_date ON video_metrics(publish_date);
CREATE INDEX IF NOT EXISTS idx_metrics_job ON video_metrics(job_id);
"""

# ── Column mapping ─────────────────────────────────────────────────────────────
_WEIXIN_MAP: dict[str, str] = {
    "视频描述": "title",
    "视频ID": "platform_id",
    "发布时间": "publish_date",
    "完播率": "completion_rate",
    "平均播放时长": "avg_watch_duration",
    "播放量": "plays",
    "推荐": "exposure",
    "喜欢": "likes",
    "评论量": "comments",
    "分享量": "shares",
    "关注量": "followers_gained",
    "转发聊天和朋友圈": "forward_count",
}

_XHS_MAP: dict[str, str] = {
    "笔记标题": "title",
    "首次发布时间": "publish_date",
    "体裁": "content_type",
    "曝光": "exposure",
    "观看量": "plays",
    "封面点击率": "cover_click_rate",
    "点赞": "likes",
    "评论": "comments",
    "收藏": "favorites",
    "涨粉": "followers_gained",
    "分享": "shares",
    "人均观看时长": "avg_watch_duration",
    "弹幕": "danmaku",
}

_INT_FIELDS = {
    "plays", "likes", "comments", "shares", "followers_gained",
    "exposure", "favorites", "danmaku", "forward_count",
}
_REAL_FIELDS = {"completion_rate", "avg_watch_duration", "cover_click_rate"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _strip_percent(val: str) -> float | None:
    """`28.64%` -> 28.64; None/empty -> None."""
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val.rstrip("%"))


def _strip_seconds(val: str) -> float | None:
    """`10.59秒` -> 10.59; None/empty -> None."""
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val.removesuffix("秒"))


def _normalize_date_weixin(val: str) -> str:
    """`2026/06/23` -> `2026-06-23`."""
    return val.strip().replace("/", "-")


def _normalize_date_xhs(val: str) -> str:
    """`2026年06月25日14时50分00秒` -> `2026-06-25`."""
    m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", str(val).strip())
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return str(val).strip()


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return int(float(val))


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val)


# ── MetricsStore ───────────────────────────────────────────────────────────────

class MetricsStore:
    """SQLite store for video metrics with CSV / XLSX import."""

    def __init__(self, db_path: str = "video_metrics.db") -> None:
        self._db_path = db_path
        self._init_db()

    # -- Connection ----------------------------------------------------------

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self._db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_db(self) -> None:
        conn = self._conn()
        try:
            conn.executescript(_SCHEMA)
            conn.commit()
        finally:
            conn.close()

    # -- Import: CSV (微信视频号) -------------------------------------------

    def import_csv(self, content: bytes, platform: str = "weixin", filename: str = "") -> dict[str, int]:
        """Import a UTF-8-BOM CSV from 微信视频号 export.

        Returns ``{"inserted": N, "updated": M}``.
        """
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        inserted = 0
        updated = 0
        conn = self._conn()
        try:
            now = datetime.now(UTC).isoformat(timespec="seconds")
            for row in reader:
                mapped = self._map_weixin_row(row, platform, now)
                if mapped is None:
                    continue
                result = self._upsert(conn, mapped)
                if result == "inserted":
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
        finally:
            conn.close()
        return {"inserted": inserted, "updated": updated}

    def _map_weixin_row(
        self, row: dict[str, str], platform: str, now: str
    ) -> dict[str, Any] | None:
        title = (row.get("视频描述") or "").strip()
        if not title:
            return None

        rec: dict[str, Any] = {"platform": platform, "imported_at": now}

        for cn_col, field in _WEIXIN_MAP.items():
            raw = row.get(cn_col)
            if raw is None:
                continue
            raw = str(raw).strip()

            if field == "publish_date":
                rec[field] = _normalize_date_weixin(raw)
            elif field in ("completion_rate", "cover_click_rate"):
                rec[field] = _strip_percent(raw)
            elif field == "avg_watch_duration":
                rec[field] = _strip_seconds(raw)
            elif field in _INT_FIELDS:
                rec[field] = _safe_int(raw)
            elif field in _REAL_FIELDS:
                rec[field] = _safe_float(raw)
            else:
                rec[field] = raw

        return rec

    # -- Import: XLSX (小红书) -----------------------------------------------

    def import_xlsx(
        self, file_path: Path, platform: str = "xiaohongshu"
    ) -> dict[str, int]:
        """Import an xlsx file in 小红书 export format.

        Skips the first ``最多导出…`` row, then reads from the ``笔记标题``
        header row onward.

        Returns ``{"inserted": N, "updated": M}``.
        """
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        inserted = 0
        updated = 0
        conn = self._conn()
        try:
            now = datetime.now(UTC).isoformat(timespec="seconds")
            header: list[str] | None = None

            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]

                # Skip preamble row(s) until we hit the header
                if header is None:
                    if any("笔记标题" in c for c in cells):
                        header = cells
                    continue

                mapped = self._map_xhs_row(cells, header, platform, now)
                if mapped is None:
                    continue
                result = self._upsert(conn, mapped)
                if result == "inserted":
                    inserted += 1
                else:
                    updated += 1

            conn.commit()
        finally:
            conn.close()
            wb.close()
        return {"inserted": inserted, "updated": updated}

    def _map_xhs_row(
        self, cells: list[str], header: list[str], platform: str, now: str
    ) -> dict[str, Any] | None:
        rec: dict[str, Any] = {"platform": platform, "imported_at": now}

        for i, col_name in enumerate(header):
            if i >= len(cells):
                break
            field = _XHS_MAP.get(col_name.strip())
            if field is None:
                continue
            raw = cells[i].strip() if cells[i] else ""
            if not raw:
                continue

            if field == "publish_date":
                rec[field] = _normalize_date_xhs(raw)
            elif field in ("completion_rate", "cover_click_rate"):
                rec[field] = _strip_percent(raw)
            elif field == "avg_watch_duration":
                rec[field] = _strip_seconds(raw)
            elif field in _INT_FIELDS:
                rec[field] = _safe_int(raw)
            elif field in _REAL_FIELDS:
                rec[field] = _safe_float(raw)
            else:
                rec[field] = raw

        if not rec.get("title"):
            return None
        return rec

    # -- Upsert helper -------------------------------------------------------

    @staticmethod
    def _upsert(conn: sqlite3.Connection, rec: dict[str, Any]) -> str:
        """INSERT or UPDATE on UNIQUE(platform, title, publish_date).

        Returns ``"inserted"`` or ``"updated"``.
        Uses a pre-check to distinguish insert from update.
        """
        existing = conn.execute(
            "SELECT 1 FROM video_metrics "
            "WHERE platform=:platform AND title=:title AND publish_date=:publish_date",
            rec,
        ).fetchone()

        if existing:
            set_parts = ", ".join(
                f"{c} = :{c}" for c in sorted(rec.keys())
                if c not in ("platform", "title", "publish_date")
            )
            conn.execute(
                f"UPDATE video_metrics SET {set_parts} "
                "WHERE platform=:platform AND title=:title AND publish_date=:publish_date",
                rec,
            )
            return "updated"

        cols = sorted(rec.keys())
        placeholders = ", ".join(f":{c}" for c in cols)
        col_names = ", ".join(cols)
        conn.execute(
            f"INSERT INTO video_metrics ({col_names}) VALUES ({placeholders})",
            rec,
        )
        return "inserted"
